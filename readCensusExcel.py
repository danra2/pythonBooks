#Read the data from the Excel spreadsheet
#Counts the number of census tracts in each county
#Counts the total population of each county.
#Prints the results
#This means your code will need to do the following

#Open and read the cells of an Excel document with the openpyxl module.
#Calculate all the tract and population data and store it in a data structure.
#Write the data structure to a text file with the .py extention using the pprint module.

import openpyxl, pprint
print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
#Opens the workbook
sheet = wb.get_sheet_by_name('Population by Census Tract')
#Selects the particular sheet that has the name.
countyData = {}

#TODO: Fill in countyData with each county's population and tracts.
print('Reading rows...')

for row in range(2, sheet.max_row + 1):
         state = sheet['B' + str(row)].value
         county = sheet['C' + str(row)].value
         pop = sheet['D' + str(row)].value

         # Make sure the key for this state exists.
         countyData.setdefault(state, {})
         # Make sure the key for this county in this state exists.
         countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
         # Each row represents one census tract, so increment by one.
         countyData[state][county]['tracts'] += 1
         # Increase the county pop by the pop in this census tract.
         countyData[state][county]['pop'] += int(pop)

#TODO: Open a new text file and write the contents of the countyData to it.

for row in range(2, sheet.max_row + 1):

# Open a new text file and write the contents of countyData to it.
    print('Writing results...')
    resultFile = open('census2010.py', 'w')
    resultFile.write('allData = ' + pprint.pformat(countyData))
    resultFile.close()
    print('Done.')
